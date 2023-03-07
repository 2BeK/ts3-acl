import csv, argparse
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

def convert():
    #print("python convert function")
    
    ## Check if Excel folder is available
    excel_path: Path = Path('excel')
    if not excel_path.is_dir():
        excel_path.mkdir()
    csv_path: Path = Path('csv')
    csv_files: Path = Path('csv').glob('**/*.csv')
    
    ## Prepare workbook
    wb= Workbook()
    
    for csv_file in csv_files:
        ## print only file name without extension
        #print(csv_file.stem)
        print("[i] Convert {} - {} data".format(csv_file.stem.split("_")[0], csv_file.stem.split("_")[1]))
        
        with open(csv_file) as csv_data:
            ## Prepare worksheet
            ws_airport_x = wb.create_sheet(str(csv_file.stem))
            ws_airport_x.title = str(csv_file.stem)

            csv_reader = csv.reader(csv_data, delimiter = ',')
            cnt = 1
            for row in csv_reader:
                #print(row)

                ws_airport_x['A'+str(cnt)] = row[0]
                ws_airport_x['B'+str(cnt)] = row[1]
                ws_airport_x['C'+str(cnt)] = row[2]
                ws_airport_x['D'+str(cnt)] = row[3]
                cnt = cnt + 1

            ## Set a default style with striped rows and baned columns
            ws_airport_style = TableStyleInfo(name="TableStyleMedium16", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            
            ## Add table information for worksheet(s)
            ws_airport_x_table = Table(displayName="Table_Airport_"+str(csv_file.stem), ref="A1:D"+str(cnt-1))
            ws_airport_x_table.tableStyleInfo = ws_airport_style
            ws_airport_x.add_table(ws_airport_x_table)
            
            ## Manual size columns
            ws_airport_x.column_dimensions['A'].width = 10
            ws_airport_x.column_dimensions['B'].width = 25
            ws_airport_x.column_dimensions['C'].width = 40
            ws_airport_x.column_dimensions['D'].width = 15
    
    del wb['Sheet']
    wb.save("excel/TS3_ACL.xlsx")
    print("[i] Finished!")


if __name__ == '__main__':
    convert()